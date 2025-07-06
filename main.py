from fastapi import FastAPI, UploadFile, File, Form, HTTPException
from fastapi.responses import FileResponse, JSONResponse, StreamingResponse
from fastapi.middleware.cors import CORSMiddleware
from openpyxl.utils.exceptions import InvalidFileException
from pydantic import BaseModel
import configparser
import os
import pickle
import zlib
import qrcode
import base64
import re
import struct
import uuid
import tempfile
import shutil
import logging
import pyzbar.pyzbar as pyzbar
import cv2
from datetime import datetime
from collections import OrderedDict
from typing import List, Dict, Any, Optional
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill, Border, Alignment
from PIL import Image, ImageDraw, ImageFont
import asyncio
import zipfile
import io

app = FastAPI()

# 允许跨域
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

# 配置
OUTPUT_DIR = "output"
os.makedirs(OUTPUT_DIR, exist_ok=True)

class QRProcessor:
    def __init__(self, output_dir):
        self.output_dir = output_dir
        os.makedirs(output_dir, exist_ok=True)

    def serialize_excel_region(self, excel_path, region, sheet_name=None, version=8, progress_callback=None):
        """序列化Excel区域"""
        if progress_callback:
            progress_callback(0, "加载Excel文件...")

        wb = load_workbook(excel_path)

        # 选择sheet
        if sheet_name:
            ws = wb[sheet_name]
        else:
            ws = wb.active

        # 解析区域坐标
        min_col, min_row, max_col, max_row = self.parse_region(region)

        # 收集数据
        data = {
            'data': [],
            'styles': [],
            'merged': [m.coord for m in ws.merged_cells.ranges],
            'meta': {
                'source': os.path.basename(excel_path),
                'sheet': sheet_name or ws.title,
                'region': region,
                'version': version,
                'timestamp': datetime.now().isoformat(),
                'mode': 'region'
            }
        }

        total_rows = max_row - min_row + 1
        for row_idx, row in enumerate(ws.iter_rows(min_row=min_row, max_row=max_row,
                                                   min_col=min_col, max_col=max_col)):
            row_data = []
            row_styles = []

            for cell in row:
                # 处理不同类型的数据
                if isinstance(cell.value, str) and len(cell.value) > 1000:
                    row_data.append(cell.value[:1000] + "...[TRUNCATED]")
                else:
                    row_data.append(cell.value)
                row_styles.append(self.get_style(cell))

            data['data'].append(row_data)
            data['styles'].append(row_styles)

            # 更新进度
            if progress_callback and total_rows > 0:
                progress = (row_idx + 1) / total_rows * 50
                progress_callback(progress, f"处理行 {row_idx + 1}/{total_rows}")

        # 序列化并压缩
        if progress_callback:
            progress_callback(60, "序列化数据...")

        serialized = pickle.dumps(data)

        if progress_callback:
            progress_callback(70, "压缩数据...")

        compressed = zlib.compress(serialized)

        # 添加校验和
        if progress_callback:
            progress_callback(80, "添加校验和...")

        checksum = zlib.crc32(compressed)
        compressed_with_checksum = struct.pack(">I", checksum) + compressed

        return compressed_with_checksum

    def serialize_file(self, file_path, version=8, progress_callback=None):
        """序列化任意文件"""
        try:
            if progress_callback:
                progress_callback(0, "读取文件...")

            # 获取文件大小
            file_size = os.path.getsize(file_path)
            if file_size > 10 * 1024 * 1024:  # 10MB
                raise ValueError("文件过大，建议使用区域模式")

            # 读取整个文件
            with open(file_path, 'rb') as f:
                file_data = f.read()

            if progress_callback:
                progress_callback(30, "压缩数据...")

            # 压缩数据
            compressed = zlib.compress(file_data)

            if progress_callback:
                progress_callback(70, "添加校验和...")

            # 添加校验和
            checksum = zlib.crc32(compressed)
            compressed_with_checksum = struct.pack(">I", checksum) + compressed

            # 添加文件模式标记
            file_marker = b"FILE_MODE:"
            final_data = file_marker + compressed_with_checksum

            # 保存副本
            filename = os.path.basename(file_path)
            with open(os.path.join(self.output_dir, f"{filename}.qrdat"), 'wb') as f:
                f.write(final_data)

            return final_data

        except Exception as e:
            raise ValueError(f"文件序列化失败: {str(e)}")

    def restore(self, data, output_path=None):
        """从数据恢复文件"""
        try:
            # 检查是否为文件模式
            if data.startswith(b"FILE_MODE:"):
                return self.restore_file(data[10:], output_path)
            else:
                return self.restore_excel_region(data, output_path)

        except Exception as e:
            # 保存原始数据用于调试
            debug_path = os.path.join(self.output_dir, "restore_debug.dat")
            with open(debug_path, 'wb') as f:
                f.write(data)
            raise ValueError(f"恢复失败: {str(e)}\n原始数据已保存至: {debug_path}")

    def restore_excel_region(self, data, output_path=None):
        """恢复Excel区域数据"""
        # 验证数据完整性
        if len(data) < 4:
            raise ValueError("数据过短，无法恢复")

        # 提取校验和
        stored_checksum = struct.unpack(">I", data[:4])[0]
        actual_data = data[4:]

        # 验证校验和
        actual_checksum = zlib.crc32(actual_data)
        if stored_checksum != actual_checksum:
            raise ValueError(f"数据校验失败: {stored_checksum} != {actual_checksum}")

        # 尝试解压数据
        try:
            decompressed = zlib.decompress(actual_data)
        except zlib.error as e:
            # 尝试不解压直接使用
            try:
                decompressed = actual_data
            except:
                raise ValueError(f"解压失败: {str(e)}")

        # 反序列化
        try:
            restored = pickle.loads(decompressed)
        except pickle.UnpicklingError as e:
            raise ValueError(f"反序列化失败: {str(e)}")

        # 检查模式
        meta = restored.get('meta', {})
        if meta.get('mode') != 'region':
            raise ValueError("数据模式不匹配，请使用文件模式恢复")

        # 检查版本兼容性
        data_version = meta.get('version', 0)
        if data_version < 4:
            raise ValueError(f"不兼容的数据版本: {data_version} (需要4+)")

        # 创建新工作簿
        wb = Workbook()
        ws = wb.active

        # 设置sheet名称
        sheet_name = meta.get('sheet', 'Restored')
        if sheet_name:
            ws.title = sheet_name[:30]  # Excel sheet名称长度限制

        # 恢复数据
        total_rows = len(restored['data'])
        for r, (row_data, row_styles) in enumerate(zip(restored['data'], restored['styles'])):
            for c, (value, style) in enumerate(zip(row_data, row_styles)):
                cell = ws.cell(row=r + 1, column=c + 1, value=value)
                self.apply_style(cell, style)

        # 恢复合并单元格
        for merged in restored.get('merged', []):
            try:
                ws.merge_cells(merged)
            except:
                continue

        # 设置输出路径
        if not output_path:
            source = meta.get('source', 'restored')
            name = os.path.splitext(source)[0]
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            output_path = os.path.join(self.output_dir, f"{name}_{timestamp}_restored.xlsx")

        wb.save(output_path)
        return output_path

    def restore_file(self, data, output_path=None):
        """恢复任意文件"""
        # 验证数据完整性
        if len(data) < 4:
            raise ValueError("数据过短，无法恢复")

        # 提取校验和
        stored_checksum = struct.unpack(">I", data[:4])[0]
        actual_data = data[4:]

        # 验证校验和
        actual_checksum = zlib.crc32(actual_data)
        if stored_checksum != actual_checksum:
            raise ValueError(f"数据校验失败: {stored_checksum} != {actual_checksum}")

        # 解压数据
        try:
            decompressed = zlib.decompress(actual_data)
        except zlib.error as e:
            raise ValueError(f"解压失败: {str(e)}")

        # 设置输出路径
        if not output_path:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            output_path = os.path.join(self.output_dir, f"restored_file_{timestamp}")

        # 保存文件
        with open(output_path, 'wb') as f:
            f.write(decompressed)

        return output_path

    def create_qr_codes(self, data, max_size=1800, version=8, mode="file", progress_callback=None):
        """生成二维码序列"""
        # 计算base64编码后的最大原始数据大小
        max_raw_size = int(max_size * 0.7)  # 考虑base64开销

        # 如果数据很小，直接生成单个二维码
        if len(data) <= max_raw_size:
            # 使用base64编码
            base64_data = base64.b64encode(data).decode('utf-8')
            return [("single", self.create_single_qr(base64_data, f"{mode}"))]

        # 计算需要多少分块
        total_chunks = (len(data) + max_raw_size - 1) // max_raw_size
        chunks = []

        # 大数据分块处理
        for i in range(total_chunks):
            start = i * max_raw_size
            end = min(start + max_raw_size, len(data))
            chunk_data = data[start:end]

            # 添加分块头并使用base64编码
            header = f"QR:{i + 1}/{total_chunks}|v{version}|{mode}|"
            base64_chunk = base64.b64encode(chunk_data).decode('utf-8')

            # 检查总长度
            full_chunk = header + base64_chunk
            if len(full_chunk) > max_size:
                # 如果超出，减小分块大小
                new_max_raw_size = int(max_raw_size * 0.9)
                return self.create_qr_codes(data, new_max_raw_size, version, mode, progress_callback)

            # 创建二维码
            name = f"chunk_{i + 1}_of_{total_chunks}"
            img = self.create_single_qr(full_chunk, f"{i + 1}/{total_chunks}")
            chunks.append((name, img))

            # 更新进度
            if progress_callback:
                progress = (i + 1) / total_chunks * 100
                progress_callback(progress, f"生成二维码 {i + 1}/{total_chunks}")

        return chunks

    def create_single_qr(self, data, counter=None):
        """创建单个二维码"""
        qr = qrcode.QRCode(
            version=None,
            error_correction=qrcode.constants.ERROR_CORRECT_L,
            box_size=10,  # 固定大小
            border=4,
        )
        qr.add_data(data)
        qr.make(fit=True)

        img = qr.make_image(fill_color="black", back_color="white")

        # 添加标记文本
        if counter:
            draw = ImageDraw.Draw(img)
            try:
                font = ImageFont.truetype("arial.ttf", 16)
            except:
                try:
                    font = ImageFont.truetype("Arial.ttf", 16)
                except:
                    font = ImageFont.load_default()

            text = counter
            bbox = draw.textbbox((0, 0), text, font=font)
            text_w = bbox[2] - bbox[0]
            text_h = bbox[3] - bbox[1]

            # 在右下角添加文本
            draw.rectangle(
                [(img.width - text_w - 10, img.height - text_h - 10),
                 (img.width, img.height)],
                fill="white"
            )
            draw.text(
                (img.width - text_w - 5, img.height - text_h - 5),
                text,
                font=font,
                fill="black"
            )

        return img

    def combine_data(self, chunks):
        """合并分块数据"""
        try:
            # 提取所有分块数据
            chunks_dict = {}
            total_chunks = 0
            current_chunks = 0
            version = 0
            mode = "file"  # 默认文件模式

            # 首先收集所有分块信息
            for chunk in chunks:
                if chunk.startswith("QR:"):
                    # 分块格式: "QR:2/5|v8|mode|base64数据"
                    parts = chunk.split('|', 3)
                    if len(parts) < 4:
                        continue

                    header = parts[0]
                    version_part = parts[1]
                    mode_part = parts[2]
                    data_part = parts[3]

                    # 解析模式
                    mode = mode_part

                    # 解析版本
                    if version_part.startswith("v"):
                        try:
                            version = int(version_part[1:])
                        except:
                            version = 0

                    # 解析分块头: "QR:2/5"
                    chunk_info = header.split(':')[1]
                    chunk_num, total = chunk_info.split('/')

                    chunks_dict[int(chunk_num)] = data_part
                    total_chunks = int(total)
                    current_chunks += 1
                else:
                    # 单个二维码情况
                    return base64.b64decode(chunk)

            # 检查是否收集到所有分块
            if current_chunks != total_chunks:
                missing = [i for i in range(1, total_chunks + 1) if i not in chunks_dict]
                raise ValueError(f"数据不完整: 缺少分块 {missing}")

            # 按顺序组合分块
            combined_b64 = ''.join(chunks_dict[i] for i in sorted(chunks_dict.keys()))
            return base64.b64decode(combined_b64)

        except Exception as e:
            # self.log(f"合并数据失败: {str(e)}")
            return None

    # 辅助方法
    def parse_region(self, region):
        """解析区域坐标 - 增强容错性"""
        # 移除空格并转换为大写
        region = region.replace(" ", "").upper()

        if ':' in region:
            start, end = region.split(':', 1)
        else:
            start = end = region

        # 使用正则表达式提取列和行
        pattern = r"([A-Z]+)(\d+)"
        start_match = re.match(pattern, start)
        end_match = re.match(pattern, end)

        if not start_match or not end_match:
            raise ValueError(f"无效的区域格式: {region}")

        start_col = start_match.group(1)
        start_row = int(start_match.group(2))
        end_col = end_match.group(1)
        end_row = int(end_match.group(2))

        # 列字母转数字
        def col_to_num(col):
            num = 0
            for c in col:
                if c.isalpha():
                    num = num * 26 + (ord(c) - ord('A')) + 1
            return num

        return (
            col_to_num(start_col),
            start_row,
            col_to_num(end_col),
            end_row
        )

    def get_style(self, cell):
        """获取单元格样式 - 增强容错性"""
        try:
            return {
                'font': self.copy_font(cell.font) if cell.font else None,
                'fill': self.copy_fill(cell.fill) if cell.fill else None,
                'border': self.copy_border(cell.border) if cell.border else None,
                'alignment': self.copy_alignment(cell.alignment) if cell.alignment else None,
                'format': cell.number_format
            }
        except Exception:
            return {}

    def apply_style(self, cell, style):
        """应用样式到单元格 - 安全处理None值"""
        if style.get('font'):
            try:
                cell.font = style['font']
            except Exception:
                pass
        if style.get('fill'):
            try:
                cell.fill = style['fill']
            except Exception:
                pass
        if style.get('border'):
            try:
                cell.border = style['border']
            except Exception:
                pass
        if style.get('alignment'):
            try:
                cell.alignment = style['alignment']
            except Exception:
                pass
        if style.get('format'):
            try:
                cell.number_format = style['format']
            except Exception:
                pass

    def copy_font(self, font):
        if not font: return None
        return Font(
            name=font.name, size=font.size, bold=font.bold,
            italic=font.italic, strike=font.strike,
            color=font.color
        )

    def copy_fill(self, fill):
        if not fill: return None
        return PatternFill(
            fill_type=fill.fill_type,
            start_color=fill.start_color,
            end_color=fill.end_color
        )

    def copy_border(self, border):
        if not border: return None
        return Border(
            left=border.left, right=border.right,
            top=border.top, bottom=border.bottom
        )

    def copy_alignment(self, alignment):
        if not alignment: return None
        return Alignment(
            horizontal=alignment.horizontal,
            vertical=alignment.vertical,
            wrap_text=alignment.wrap_text,
            shrink_to_fit=alignment.shrink_to_fit,
            indent=alignment.indent
        )

# class QRProcessor:
#     # 保持原 QRProcessor 类实现不变
#     # 仅修改文件路径处理方式
#     pass


class SerializeRequest(BaseModel):
    mode: str  # "region" 或 "file"
    file_path: Optional[str] = None
    region: Optional[str] = None
    sheet_name: Optional[str] = None
    version: int = 8
    max_chunk_size: int = 1800


class QRGenerationRequest(BaseModel):
    session_id: str
    max_chunk_size: int = 1800


class ScanRequest(BaseModel):
    # session_id: str
    files: List[str]  # base64 编码的图片列表


class VideoScanRequest(BaseModel):
    session_id: str
    video: str  # base64 编码的视频文件


# 会话状态存储
sessions = {}


@app.post("/serialize")
async def serialize_data(request: SerializeRequest):
    session_id = str(uuid.uuid4())
    sessions[session_id] = {
        "status": "processing",
        "progress": 0,
        "message": "开始序列化...",
        "serialized_data": None,
        "mode": request.mode,
        "file_path": None,
        "version": request.version
    }

    try:
        # 处理文件上传
        file_data = request.file_path
        if not file_data:
            raise HTTPException(status_code=400, detail="未提供文件数据")

        # 保存文件到临时目录
        file_ext = ".xlsx" if request.mode == "region" else ".bin"
        file_path = os.path.join(OUTPUT_DIR, f"{session_id}_source{file_ext}")
        with open(file_path, "wb") as f:
            f.write(base64.b64decode(file_data.split(",")[1]))

        processor = QRProcessor(OUTPUT_DIR)

        if request.mode == "region":
            sessions[session_id]["message"] = "序列化Excel区域..."
            serialized_data = processor.serialize_excel_region(
                file_path,
                request.region or "A1:D10",
                sheet_name=request.sheet_name,
                version=request.version
            )
        else:
            sessions[session_id]["message"] = "序列化文件..."
            serialized_data = processor.serialize_file(
                file_path,
                version=request.version
            )

        sessions[session_id].update({
            "status": "completed",
            "progress": 100,
            "message": "序列化完成",
            "serialized_data": base64.b64encode(serialized_data).decode(),
            "file_path": file_path
        })

        return JSONResponse(content={
            "session_id": session_id,
            "data_size": len(serialized_data),
            "message": "序列化成功"
        })

    except Exception as e:
        sessions[session_id].update({
            "status": "error",
            "message": f"序列化失败: {str(e)}"
        })
        raise HTTPException(status_code=500, detail=str(e))


@app.post("/generate-qr")
async def generate_qr(request: QRGenerationRequest):
    session = sessions.get(request.session_id)
    if not session:
        raise HTTPException(status_code=404, detail="会话不存在")

    if session["status"] != "completed" or not session.get("serialized_data"):
        raise HTTPException(status_code=400, detail="请先完成序列化")

    try:
        session["status"] = "processing"
        session["progress"] = 0
        session["message"] = "开始生成二维码..."

        serialized_data = base64.b64decode(session["serialized_data"])
        processor = QRProcessor(OUTPUT_DIR)

        # 生成二维码
        qr_images = processor.create_qr_codes(
            serialized_data,
            max_size=request.max_chunk_size,
            version=session["version"],
            mode=session["mode"]
        )

        # 保存二维码图片
        qr_files = []
        for i, (name, img) in enumerate(qr_images):
            img_path = os.path.join(OUTPUT_DIR, f"{request.session_id}_qr_{i}.png")
            img.save(img_path)
            qr_files.append({
                "name": name,
                "path": img_path
            })

            session["progress"] = int((i + 1) / len(qr_images) * 100)
            session["message"] = f"生成二维码 {i + 1}/{len(qr_images)}"
            await asyncio.sleep(0.01)  # 让出控制权

        session["qr_images"] = qr_files
        session["status"] = "completed"
        session["message"] = "二维码生成完成"

        # 返回二维码预览
        previews = []
        for qr in qr_files[:3]:  # 只返回前3个预览
            with open(qr["path"], "rb") as f:
                previews.append({
                    "name": qr["name"],
                    "data": base64.b64encode(f.read()).decode()
                })

        return JSONResponse(content={
            "count": len(qr_files),
            "previews": previews,
            "message": f"成功生成 {len(qr_files)} 个二维码"
        })

    except Exception as e:
        session["status"] = "error"
        session["message"] = f"生成失败: {str(e)}"
        raise HTTPException(status_code=500, detail=str(e))


from fastapi import UploadFile, File
from openpyxl import load_workbook
import tempfile
import os


@app.post("/get-sheets")
async def get_sheets(file: UploadFile = File(...)):
    # 验证文件扩展名
    if not file.filename.lower().endswith(('.xlsx', '.xlsm', '.xltx', '.xltm')):
        raise HTTPException(
            status_code=400,
            detail="仅支持 .xlsx, .xlsm, .xltx, .xltm 格式的Excel文件"
        )

    try:
        # 方法1：直接从内存读取（推荐）
        contents = await file.read()
        try:
            wb = load_workbook(io.BytesIO(contents))
            return {"sheets": wb.sheetnames}
        except InvalidFileException:
            # 方法2：如果内存读取失败，尝试临时文件方式
            with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
                tmp.write(contents)
                tmp_path = tmp.name

            try:
                wb = load_workbook(tmp_path)
                return {"sheets": wb.sheetnames}
            finally:
                os.unlink(tmp_path)

    except Exception as e:
        raise HTTPException(
            status_code=500,
            detail=f"读取Excel失败: {str(e)}"
        )


@app.get("/session/{session_id}")
async def get_session_status(session_id: str):
    session = sessions.get(session_id)
    if not session:
        raise HTTPException(status_code=404, detail="会话不存在")

    return JSONResponse(content={
        "status": session["status"],
        "progress": session["progress"],
        "message": session["message"]
    })


@app.post("/scan-images")
async def scan_images(request: ScanRequest):
    session_id = str(uuid.uuid4())
    sessions[session_id] = {
        "status": "processing",
        "progress": 0,
        "message": "开始扫描二维码...",
        "restored_file": None
    }

    try:
        chunks = []
        total_files = len(request.files)

        for i, file_data in enumerate(request.files):
            # 解码base64图片
            img_data = base64.b64decode(file_data.split(",")[1])
            img = Image.open(io.BytesIO(img_data))

            # 解码二维码
            results = pyzbar.decode(img)
            for r in results:
                if r.type == 'QRCODE':
                    chunks.append(r.data.decode('utf-8'))

            sessions[session_id]["progress"] = int((i + 1) / total_files * 100)
            sessions[session_id]["message"] = f"扫描文件 {i + 1}/{total_files}"
            await asyncio.sleep(0.01)

        if not chunks:
            raise ValueError("未找到有效二维码数据")

        # 合并数据
        sessions[session_id]["message"] = "合并数据..."
        processor = QRProcessor(OUTPUT_DIR)
        combined = processor.combine_data(chunks)

        if not combined:
            raise ValueError("数据不完整")

        # 恢复文件
        sessions[session_id]["message"] = "恢复文件..."
        output_path = processor.restore(combined)

        # 读取恢复的文件
        with open(output_path, "rb") as f:
            file_content = f.read()

        sessions[session_id].update({
            "status": "completed",
            "progress": 100,
            "message": "恢复完成",
            "restored_file": base64.b64encode(file_content).decode(),
            "file_name": os.path.basename(output_path)
        })

        return JSONResponse(content={
            "file_name": os.path.basename(output_path),
            "message": "文件恢复成功"
        })

    except Exception as e:
        sessions[session_id].update({
            "status": "error",
            "message": f"恢复失败: {str(e)}"
        })
        raise HTTPException(status_code=500, detail=str(e))


@app.post("/scan-video")
async def scan_video(request: VideoScanRequest):
    session_id = str(uuid.uuid4())
    sessions[session_id] = {
        "status": "processing",
        "progress": 0,
        "message": "开始扫描视频...",
        "restored_file": None
    }

    try:
        # 保存视频到临时文件
        video_data = base64.b64decode(request.video.split(",")[1])
        video_path = os.path.join(OUTPUT_DIR, f"{session_id}_video.mp4")
        with open(video_path, "wb") as f:
            f.write(video_data)

        # 扫描视频
        unique_qrs = OrderedDict()

        cap = cv2.VideoCapture(video_path)
        if not cap.isOpened():
            raise ValueError("无法打开视频文件")

        frame_count = int(cap.get(cv2.CAP_PROP_FRAME_COUNT))
        scanned_frames = 0
        unique_count = 0

        while cap.isOpened():
            ret, frame = cap.read()
            if not ret:
                break

            scanned_frames += 1
            progress = min(100, int((scanned_frames / frame_count) * 100))

            # 更新进度
            sessions[session_id]["progress"] = progress
            sessions[session_id]["message"] = f"扫描中... ({scanned_frames}/{frame_count} 帧)"
            await asyncio.sleep(0.01)  # 让出控制权

            # 解码二维码
            frame_rgb = cv2.cvtColor(frame, cv2.COLOR_BGR2RGB)
            pil_image = Image.fromarray(frame_rgb)
            decoded_objects = pyzbar.decode(pil_image)

            if decoded_objects:
                for obj in decoded_objects:
                    if obj.type == 'QRCODE':
                        qr_data = obj.data.decode('utf-8')
                        if qr_data not in unique_qrs:
                            unique_qrs[qr_data] = True
                            unique_count += 1
                            sessions[session_id]["message"] = f"发现新二维码: #{unique_count}"

        cap.release()

        if not unique_qrs:
            raise ValueError("未在视频中发现二维码")

        # 合并数据
        sessions[session_id]["message"] = "合并数据..."
        processor = QRProcessor(OUTPUT_DIR)
        combined = processor.combine_data(list(unique_qrs.keys()))

        if not combined:
            raise ValueError("数据不完整")

        # 恢复文件
        sessions[session_id]["message"] = "恢复文件..."
        output_path = processor.restore(combined)

        # 读取恢复的文件
        with open(output_path, "rb") as f:
            file_content = f.read()

        sessions[session_id].update({
            "status": "completed",
            "progress": 100,
            "message": "恢复完成",
            "restored_file": base64.b64encode(file_content).decode(),
            "file_name": os.path.basename(output_path)
        })

        return JSONResponse(content={
            "file_name": os.path.basename(output_path),
            "message": "文件恢复成功"
        })

    except Exception as e:
        sessions[session_id].update({
            "status": "error",
            "message": f"视频恢复失败: {str(e)}"
        })
        raise HTTPException(status_code=500, detail=str(e))


@app.get("/download/{session_id}")
async def download_files(session_id: str, file_type: str):
    session = sessions.get(session_id)
    if not session:
        raise HTTPException(status_code=404, detail="会话不存在")

    try:
        if file_type == "qr" and session.get("qr_images"):
            # 创建ZIP文件
            zip_buffer = io.BytesIO()
            with zipfile.ZipFile(zip_buffer, "w") as zipf:
                for qr in session["qr_images"]:
                    zipf.write(qr["path"], os.path.basename(qr["path"]))

            zip_buffer.seek(0)
            return StreamingResponse(
                zip_buffer,
                media_type="application/zip",
                headers={"Content-Disposition": f"attachment; filename=qr_codes_{session_id}.zip"}
            )

        elif file_type == "restored" and session.get("restored_file"):
            file_content = base64.b64decode(session["restored_file"])
            return StreamingResponse(
                io.BytesIO(file_content),
                media_type="application/octet-stream",
                headers={"Content-Disposition": f"attachment; filename={session['file_name']}"}
            )

        else:
            raise HTTPException(status_code=404, detail="文件不存在")

    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


# 配置管理端点
@app.get("/config")
async def get_config():
    config_path = os.path.join(OUTPUT_DIR, "config.ini")
    if not os.path.exists(config_path):
        return JSONResponse(content={"config": {}})

    config = configparser.ConfigParser()
    config.read(config_path)
    return JSONResponse(content=dict(config))


@app.post("/config")
async def save_config(config_data: dict):
    try:
        config = configparser.ConfigParser()
        config.read_dict(config_data)

        config_path = os.path.join(OUTPUT_DIR, "config.ini")
        with open(config_path, "w") as configfile:
            config.write(configfile)

        return JSONResponse(content={"message": "配置保存成功"})
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.post("/reset-config")
async def reset_config():
    try:
        config_path = os.path.join(OUTPUT_DIR, "config.ini")
        if os.path.exists(config_path):
            os.remove(config_path)
        return JSONResponse(content={"message": "配置已重置"})
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


if __name__ == '__main__':
    import uvicorn
    uvicorn.run(app,host='0.0.0.0',port=8000)
import configparser
import os
import pickle
import zlib
import qrcode
import tkinter as tk
from tkinter import filedialog, messagebox, ttk, scrolledtext
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill, Border, Alignment
from PIL import Image, ImageTk, ImageDraw, ImageFont
import threading
import logging
import pyzbar.pyzbar as pyzbar
import base64
import re
from datetime import datetime
import struct
import time
import cv2  # 用于视频处理
from collections import OrderedDict


class VideoQRScanner:
    """视频二维码扫描器"""

    def __init__(self, video_path, output_dir, callback):
        self.video_path = video_path
        self.output_dir = output_dir
        self.callback = callback
        self.cap = None
        self.running = False
        self.unique_qrs = OrderedDict()  # 使用有序字典存储唯一二维码
        self.current_frame = None
        self.frame_count = 0
        self.scanned_frames = 0
        self.unique_count = 0
        self.last_qr_data = None
        self.scan_interval = 0.01  # 帧处理间隔（秒）
        self.min_confidence = 30  # 最小置信度阈值

    def start(self):
        """开始扫描视频"""
        try:
            self.cap = cv2.VideoCapture(self.video_path)
            if not self.cap.isOpened():
                raise ValueError("无法打开视频文件")

            self.running = True
            self.frame_count = int(self.cap.get(cv2.CAP_PROP_FRAME_COUNT))
            self.callback(0, f"开始扫描视频: {os.path.basename(self.video_path)}")
            self.process_video()
        except Exception as e:
            self.callback(0, f"视频扫描错误: {str(e)}")
            self.stop()

    def process_video(self):
        """处理视频帧"""
        if not self.running or not self.cap.isOpened():
            return

        ret, frame = self.cap.read()
        if not ret:
            self.stop()
            self.callback(100, "视频处理完成")
            return

        self.scanned_frames += 1
        progress = min(100, int((self.scanned_frames / self.frame_count) * 100))
        self.callback(progress, f"扫描中... ({self.scanned_frames}/{self.frame_count} 帧)")

        # 转换为PIL图像进行处理
        frame_rgb = cv2.cvtColor(frame, cv2.COLOR_BGR2RGB)
        pil_image = Image.fromarray(frame_rgb)
        self.current_frame = pil_image

        # 解码二维码
        decoded_objects = pyzbar.decode(pil_image)

        if decoded_objects:
            for obj in decoded_objects:
                if obj.type == 'QRCODE':
                    qr_data = obj.data.decode('utf-8')

                    # 检查是否是新二维码
                    if qr_data not in self.unique_qrs:
                        self.unique_qrs[qr_data] = pil_image
                        self.unique_count += 1
                        self.last_qr_data = qr_data
                        self.callback(progress, f"发现新二维码: #{self.unique_count}")

        # 继续处理下一帧
        threading.Timer(self.scan_interval, self.process_video).start()

    def stop(self):
        """停止扫描"""
        self.running = False
        if self.cap and self.cap.isOpened():
            self.cap.release()

    def get_results(self):
        """获取扫描结果"""
        return self.unique_qrs


class FileQRApp:
    def __init__(self, root):
        self.root = root
        self.root.title("文件二维码工具 v8.2")
        self.root.geometry("1350x900")

        # 设置应用图标
        try:
            self.root.iconbitmap("icon.ico")
        except:
            pass

        # 初始化
        self.output_dir = "output"
        os.makedirs(self.output_dir, exist_ok=True)
        self.setup_logging()
        self.init_variables()
        self.create_ui()
        # 加载配置
        self.load_config()

        self.log("应用程序启动")

    def setup_logging(self):
        """配置日志系统"""
        logging.basicConfig(
            filename=os.path.join(self.output_dir, 'app.log'),
            level=logging.INFO,
            format='%(asctime)s - %(levelname)s - %(message)s'
        )

    def log(self, message):
        """记录日志"""
        logging.info(message)
        self.status.set(message)
        self.add_log_entry(message)

    def init_variables(self):
        """初始化变量"""
        self.serialized_data = None
        self.qr_images = []
        self.current_qr_index = 0
        self.max_chunk_size = 1800  # 数据块大小
        self.version = 8  # 当前版本号
        self.mode = "region"  # 默认模式：Excel区域模式
        self.progress_value = 0
        self.log_entries = []
        self.log_visible = True
        self.last_region = "A1:D10"  # 默认区域
        self.last_sheet = ""  # 默认Sheet
        self.video_scanner = None  # 视频扫描器实例

    def create_ui(self):
        """创建用户界面"""
        # 创建主框架
        self.main_frame = ttk.Frame(self.root, padding="10")
        self.main_frame.pack(fill=tk.BOTH, expand=True)

        # 创建左侧控制面板
        self.create_control_panel()

        # 创建右侧日志面板
        self.create_log_panel()

        # 创建二维码显示区域
        self.create_qr_display()

        self.bind_events()

    def create_control_panel(self):
        """创建左侧控制面板"""
        control_frame = ttk.LabelFrame(self.main_frame, text="控制面板")
        control_frame.pack(side=tk.LEFT, fill=tk.Y, padx=5, pady=5, ipadx=5, ipady=5)

        # 模式选择
        mode_frame = ttk.LabelFrame(control_frame, text="模式选择")
        mode_frame.pack(fill=tk.X, pady=5)

        self.mode_var = tk.StringVar(value="region")
        ttk.Radiobutton(mode_frame, text="Excel区域模式", variable=self.mode_var,
                        value="region", command=self.toggle_mode).pack(side=tk.LEFT, padx=10)
        ttk.Radiobutton(mode_frame, text="文件模式", variable=self.mode_var,
                        value="file", command=self.toggle_mode).pack(side=tk.LEFT, padx=10)

        # 文件选择
        file_frame = ttk.LabelFrame(control_frame, text="选择文件")
        file_frame.pack(fill=tk.X, pady=5)

        self.file_path = tk.StringVar()
        file_entry = ttk.Entry(file_frame, textvariable=self.file_path, width=30)
        file_entry.pack(side=tk.LEFT, padx=5, fill=tk.X, expand=True)
        ttk.Button(file_frame, text="浏览", command=self.browse_file, width=8).pack(side=tk.LEFT)

        # Excel区域设置
        self.region_frame = ttk.LabelFrame(control_frame, text="Excel区域设置")
        self.region_frame.pack(fill=tk.X, pady=5)

        ttk.Label(self.region_frame, text="区域:").pack(side=tk.LEFT, padx=5)
        self.region_entry = ttk.Entry(self.region_frame, width=20)
        self.region_entry.pack(side=tk.LEFT, padx=5, fill=tk.X, expand=True)
        self.region_entry.insert(0, self.last_region)

        # Sheet选择
        self.sheet_frame = ttk.Frame(control_frame)
        self.sheet_frame.pack(fill=tk.X, pady=5)

        ttk.Label(self.sheet_frame, text="选择Sheet:").pack(side=tk.LEFT, padx=5)
        self.sheet_var = tk.StringVar(value=self.last_sheet)
        self.sheet_combo = ttk.Combobox(self.sheet_frame, textvariable=self.sheet_var,
                                        state="readonly", width=15)
        self.sheet_combo.pack(side=tk.LEFT, padx=5, fill=tk.X, expand=True)

        # 二维码设置
        qr_frame = ttk.LabelFrame(control_frame, text="二维码设置")
        qr_frame.pack(fill=tk.X, pady=5)

        # 容量设置
        capacity_frame = ttk.Frame(qr_frame)
        capacity_frame.pack(fill=tk.X, pady=3)

        ttk.Label(capacity_frame, text="数据块大小:").pack(side=tk.LEFT, padx=5)
        self.capacity_var = tk.StringVar(value=str(self.max_chunk_size))
        capacity_combo = ttk.Combobox(capacity_frame, textvariable=self.capacity_var,
                                      values=["1200", "1500", "1800", "2000", "2500"], width=8)
        capacity_combo.pack(side=tk.LEFT, padx=5)
        ttk.Label(capacity_frame, text="字节").pack(side=tk.LEFT)

        # 版本设置
        version_frame = ttk.Frame(qr_frame)
        version_frame.pack(fill=tk.X, pady=3)

        ttk.Label(version_frame, text="数据版本:").pack(side=tk.LEFT, padx=5)
        self.version_var = tk.StringVar(value=str(self.version))
        ttk.Label(version_frame, textvariable=self.version_var).pack(side=tk.LEFT, padx=5)

        # 进度条
        self.progress_frame = ttk.LabelFrame(control_frame, text="进度")
        self.progress_frame.pack(fill=tk.X, pady=5)

        self.progress_var = tk.DoubleVar()
        self.progress_bar = ttk.Progressbar(self.progress_frame, variable=self.progress_var,
                                            maximum=100, length=300)
        self.progress_bar.pack(fill=tk.X, padx=5, pady=5)

        self.progress_label = ttk.Label(self.progress_frame, text="就绪")
        self.progress_label.pack(pady=5)

        # 操作按钮
        btn_frame = ttk.Frame(control_frame)
        btn_frame.pack(fill=tk.X, pady=10)

        ttk.Button(btn_frame, text="序列化", command=self.serialize).pack(side=tk.LEFT, padx=5, fill=tk.X, expand=True)
        ttk.Button(btn_frame, text="生成二维码", command=self.generate_qr).pack(side=tk.LEFT, padx=5, fill=tk.X, expand=True)
        ttk.Button(btn_frame, text="图片恢复", command=self.scan_restore).pack(side=tk.LEFT, padx=5, fill=tk.X, expand=True)
        ttk.Button(btn_frame, text="视频恢复", command=self.scan_video).pack(side=tk.LEFT, padx=5, fill=tk.X, expand=True)
        ttk.Button(btn_frame, text="保存二维码", command=self.save_qr_images).pack(side=tk.LEFT, padx=5, fill=tk.X,
                                                                              expand=True)

        # 日志控制按钮
        log_control_frame = ttk.Frame(control_frame)
        log_control_frame.pack(fill=tk.X, pady=5)

        ttk.Button(log_control_frame, text="清空日志", command=self.clear_logs).pack(side=tk.LEFT, padx=5)

        # 修复：将按钮保存为实例变量
        self.toggle_log_btn = ttk.Button(log_control_frame, text="隐藏日志", command=self.toggle_log_visibility)
        self.toggle_log_btn.pack(side=tk.RIGHT, padx=5)


        # 配置按钮
        config_frame = ttk.Frame(control_frame)
        config_frame.pack(fill=tk.X, pady=5)
        #
        ttk.Button(config_frame, text="保存配置", command=self.save_config).pack(side=tk.LEFT, padx=5, fill=tk.X,
                                                                             expand=True)
        ttk.Button(config_frame, text="重置配置", command=self.reset_config).pack(side=tk.LEFT, padx=5, fill=tk.X,
                                                                              expand=True)

    def load_config(self):
        """加载配置文件 - 修复版"""
        config_path = os.path.join(self.output_dir, "config.ini")
        if not os.path.exists(config_path):
            self.log("配置文件不存在，使用默认配置")
            return

        try:
            config = configparser.ConfigParser()
            config.read(config_path)

            # 加载通用设置
            if 'General' in config:
                max_chunk = config['General'].getint('max_chunk_size', self.max_chunk_size)
                if max_chunk >= 1000:  # 有效性检查
                    self.max_chunk_size = max_chunk
                    self.capacity_var.set(str(self.max_chunk_size))

            # 加载区域模式设置
            if 'RegionMode' in config:
                region = config['RegionMode'].get('last_region', 'A1:D10')
                sheet = config['RegionMode'].get('last_sheet', '')

                # 更新UI
                self.region_entry.delete(0, tk.END)
                self.region_entry.insert(0, region)
                self.sheet_var.set(sheet)

                # 更新内部变量
                self.last_region = region
                self.last_sheet = sheet

            self.log("配置加载成功")
        except Exception as e:
            self.log(f"加载配置失败: {str(e)}")

    def scan_video(self):
        """从视频中扫描二维码"""
        file_path = filedialog.askopenfilename(
            title="选择视频文件",
            filetypes=[("视频文件", "*.mp4 *.avi *.mov *.mkv")]
        )

        if not file_path:
            return

        def task():
            try:
                self.update_progress(0, "开始扫描视频...")
                self.video_scanner = VideoQRScanner(file_path, self.output_dir, self.update_progress)
                self.video_scanner.start()

                # 等待扫描完成
                while self.video_scanner and self.video_scanner.running:
                    time.sleep(0.5)

                if self.video_scanner:
                    qr_data = self.video_scanner.get_results()
                    if not qr_data:
                        self.log("未在视频中发现二维码")
                        messagebox.showinfo("提示", "未在视频中发现二维码")
                        return

                    self.log(f"在视频中发现 {len(qr_data)} 个唯一二维码")
                    self.update_progress(70, "合并数据...")

                    # 合并数据
                    combined = self.combine_data(list(qr_data.keys()))
                    if not combined:
                        raise ValueError("数据不完整")

                    # 恢复数据
                    self.update_progress(80, "恢复文件...")
                    processor = QRProcessor(self.output_dir)
                    output_path = processor.restore(combined)

                    self.log(f"文件已恢复至: {output_path}")
                    self.update_progress(100, "恢复完成！")
                    messagebox.showinfo("成功", f"文件已恢复至:\n{output_path}")
            except Exception as e:
                self.log(f"视频恢复失败: {str(e)}")
                self.update_progress(0, f"视频恢复失败: {str(e)}")
                messagebox.showerror("错误", f"视频恢复失败: {str(e)}")
                logging.exception("视频恢复失败")
            finally:
                self.update_progress(0, "就绪")
                self.video_scanner = None

        threading.Thread(target=task, daemon=True).start()

    def reset_config(self):
        """重置配置"""
        self.max_chunk_size = 1800
        self.capacity_var.set(str(self.max_chunk_size))
        self.region_entry.delete(0, tk.END)
        self.region_entry.insert(0, "A1:D10")
        self.sheet_var.set("")
        self.log("配置已重置为默认值")

    def save_config(self):
        """保存配置文件 - 增强错误处理"""
        try:
            config = configparser.ConfigParser()

            # 通用设置
            config['General'] = {
                'max_chunk_size': str(self.max_chunk_size)
            }

            # 区域模式设置
            region = self.region_entry.get() or self.last_region
            sheet = self.sheet_var.get() or self.last_sheet

            config['RegionMode'] = {
                'last_region': region,
                'last_sheet': sheet
            }

            # 保存到文件
            config_path = os.path.join(self.output_dir, "config.ini")
            with open(config_path, 'w') as configfile:
                config.write(configfile)

            # 更新内部状态
            self.last_region = region
            self.last_sheet = sheet

            self.log("配置保存成功")
            messagebox.showinfo("成功", "配置已保存")
        except Exception as e:
            self.log(f"保存配置失败: {str(e)}")
            messagebox.showerror("错误", f"保存配置失败: {str(e)}")

    def create_log_panel(self):
        """创建右侧日志面板"""
        self.log_frame = ttk.LabelFrame(self.main_frame, text="操作记录", width=250)
        self.log_frame.pack(side=tk.RIGHT, fill=tk.BOTH, padx=5, pady=5, expand=True, ipadx=5, ipady=5)


        # 日志文本框 - 设置宽度为30个字符（约300像素）
        self.log_text = scrolledtext.ScrolledText(self.log_frame, wrap=tk.WORD, height=15, width=30)
        self.log_text.pack(fill=tk.BOTH, expand=True, padx=5, pady=5)
        self.log_text.config(state=tk.DISABLED)

        # 添加欢迎消息
        self.add_log_entry("=== 文件二维码工具 v8.2 ===")
        self.add_log_entry("就绪，请选择操作")

    def create_qr_display(self):
        """创建二维码显示区域"""
        display_frame = ttk.LabelFrame(self.main_frame, text="二维码预览")
        display_frame.pack(fill=tk.BOTH, expand=True, padx=5, pady=5)

        # 导航按钮
        nav_frame = ttk.Frame(display_frame)
        nav_frame.pack(fill=tk.X, pady=5)

        self.prev_btn = ttk.Button(nav_frame, text="上一个", command=self.prev_qr, state=tk.DISABLED)
        self.prev_btn.pack(side=tk.LEFT, padx=5)

        self.next_btn = ttk.Button(nav_frame, text="下一个", command=self.next_qr, state=tk.DISABLED)
        self.next_btn.pack(side=tk.LEFT, padx=5)

        self.qr_pos_label = ttk.Label(nav_frame, text="0/0")
        self.qr_pos_label.pack(side=tk.LEFT, padx=5)

        self.qr_info_label = ttk.Label(nav_frame, text="")
        self.qr_info_label.pack(side=tk.LEFT, padx=10)

        ttk.Label(nav_frame, text="提示：点击'保存二维码'导出所有图片").pack(side=tk.RIGHT, padx=10)

        # 二维码画布
        self.qr_canvas = tk.Canvas(display_frame, bg='white', highlightthickness=0, width=800)
        self.qr_canvas.pack(fill=tk.BOTH, expand=True)

        # 状态栏
        self.status = tk.StringVar(value="就绪")
        ttk.Label(self.main_frame, textvariable=self.status).pack(fill=tk.X, pady=5)

    def bind_events(self):
        """绑定事件"""
        self.qr_canvas.bind("<Configure>", self.resize_qr)
        self.capacity_var.trace("w", self.update_capacity)
        self.file_path.trace("w", self.file_changed)

    def toggle_mode(self):
        """切换模式"""
        self.mode = self.mode_var.get()

        # 重置数据
        self.serialized_data = None
        self.qr_images = []
        self.current_qr_index = 0
        self.show_qr()

        if self.mode == "region":
            self.region_frame.pack(fill=tk.X, pady=5)
            self.sheet_frame.pack(fill=tk.X, pady=5)
        else:  # 文件模式
            self.region_frame.pack_forget()
            self.sheet_frame.pack_forget()

    def toggle_log_visibility(self):
        """切换日志可见性"""
        self.log_visible = not self.log_visible
        if self.log_visible:
            self.log_frame.pack(side=tk.RIGHT, fill=tk.BOTH, padx=5, pady=5, expand=True, ipadx=5, ipady=5)
            self.toggle_log_btn.config(text="隐藏日志")
        else:
            self.log_frame.pack_forget()
            self.toggle_log_btn.config(text="显示日志")

    def add_log_entry(self, message):
        """添加日志条目"""
        timestamp = datetime.now().strftime("%H:%M:%S")
        log_entry = f"[{timestamp}] {message}"
        self.log_entries.append(log_entry)

        if self.log_text:
            self.log_text.config(state=tk.NORMAL)
            self.log_text.insert(tk.END, log_entry + "\n")
            self.log_text.see(tk.END)
            self.log_text.config(state=tk.DISABLED)

    def clear_logs(self):
        """清空日志"""
        self.log_entries = []
        if self.log_text:
            self.log_text.config(state=tk.NORMAL)
            self.log_text.delete(1.0, tk.END)
            self.log_text.config(state=tk.DISABLED)
            self.add_log_entry("日志已清空")

    def update_progress(self, value, message=None):
        """更新进度条"""
        self.progress_var.set(value)
        if message:
            self.progress_label.config(text=message)
        self.root.update_idletasks()

    # 核心功能方法
    def serialize(self):
        """序列化数据"""
        if not self.validate_inputs():
            return

        def task():
            try:
                self.update_progress(0, "开始序列化...")
                time.sleep(0.1)  # 让UI更新
                processor = QRProcessor(self.output_dir)
                file_path = self.file_path.get()

                if self.mode == "region":
                    region = self.region_entry.get()
                    sheet_name = self.sheet_var.get()
                    self.serialized_data = processor.serialize_excel_region(
                        file_path,
                        region,
                        sheet_name=sheet_name,
                        version=self.version,
                        progress_callback=self.update_progress
                    )
                    data_size = len(self.serialized_data)
                    self.log(f"Excel区域序列化完成，数据大小: {data_size} 字节")
                    self.update_progress(100, "序列化完成！")
                    messagebox.showinfo("成功", f"Excel区域序列化成功！数据大小: {data_size} 字节")
                else:  # 文件模式
                    self.serialized_data = processor.serialize_file(
                        file_path,
                        version=self.version,
                        progress_callback=self.update_progress
                    )
                    data_size = len(self.serialized_data)
                    self.log(f"文件序列化完成，数据大小: {data_size} 字节")
                    self.update_progress(100, "序列化完成！")
                    messagebox.showinfo("成功", f"文件序列化成功！数据大小: {data_size} 字节")

            except Exception as e:
                self.log(f"序列化失败: {str(e)}")
                self.update_progress(0, f"序列化失败: {str(e)}")
                messagebox.showerror("错误", f"序列化失败: {str(e)}")
                logging.exception("序列化失败")
            finally:
                self.update_progress(0, "就绪")

        threading.Thread(target=task, daemon=True).start()

    def generate_qr(self):
        """生成二维码"""
        if not self.serialized_data:
            messagebox.showwarning("警告", "请先序列化数据")
            return

        def task():
            try:
                self.update_progress(0, "开始生成二维码...")
                time.sleep(0.1)  # 让UI更新
                processor = QRProcessor(self.output_dir)
                chunk_size = int(self.capacity_var.get())

                # 生成二维码
                self.qr_images = processor.create_qr_codes(
                    self.serialized_data,
                    max_size=chunk_size,
                    version=self.version,
                    mode=self.mode,
                    progress_callback=self.update_progress
                )

                # 显示第一个二维码
                self.current_qr_index = 0
                self.update_navigation()
                self.show_qr()

                self.log(f"成功生成 {len(self.qr_images)} 个二维码")
                self.update_progress(100, "二维码生成完成！")
                messagebox.showinfo("成功", f"二维码生成完成！共 {len(self.qr_images)} 个二维码")
            except Exception as e:
                self.log(f"生成失败: {str(e)}")
                self.update_progress(0, f"生成失败: {str(e)}")
                messagebox.showerror("错误", f"生成失败: {str(e)}")
                logging.exception("生成失败")
            finally:
                self.update_progress(0, "就绪")

        threading.Thread(target=task, daemon=True).start()

    def scan_restore(self):
        """扫描二维码图片恢复数据"""
        files = filedialog.askopenfilenames(
            title="选择二维码图片",
            filetypes=[("图片文件", "*.png *.jpg *.jpeg *.bmp")]
        )

        if not files:
            return

        def task():
            try:
                self.update_progress(0, "开始扫描恢复...")
                time.sleep(0.1)  # 让UI更新
                processor = QRProcessor(self.output_dir)

                # 解码二维码
                chunks = []
                total_files = len(files)
                for i, f in enumerate(files):
                    self.update_progress(i / total_files * 50, f"扫描文件 {i + 1}/{total_files}")
                    decoded = self.decode_qr(f)
                    if decoded:
                        chunks.extend(decoded)
                    time.sleep(0.05)  # 避免UI卡顿

                if not chunks:
                    raise ValueError("未找到有效二维码数据")

                # 合并数据
                self.update_progress(60, "合并数据...")
                combined = self.combine_data(chunks)
                if not combined:
                    raise ValueError("数据不完整")

                # 恢复数据
                self.update_progress(70, "恢复文件...")
                output_path = processor.restore(combined)

                self.log(f"文件已恢复至: {output_path}")
                self.update_progress(100, "恢复完成！")
                messagebox.showinfo("成功", f"文件已恢复至:\n{output_path}")
            except Exception as e:
                self.log(f"恢复失败: {str(e)}")
                self.update_progress(0, f"恢复失败: {str(e)}")
                messagebox.showerror("错误", f"恢复失败: {str(e)}")
                logging.exception("恢复失败")
            finally:
                self.update_progress(0, "就绪")

        threading.Thread(target=task, daemon=True).start()

    def save_qr_images(self):
        """保存所有二维码图片"""
        if not self.qr_images:
            messagebox.showwarning("警告", "没有可保存的二维码")
            return

        dir_path = filedialog.askdirectory(title="选择保存位置")
        if not dir_path:
            return

        try:
            total = len(self.qr_images)
            for idx, (name, img) in enumerate(self.qr_images):
                self.update_progress(idx / total * 100, f"保存二维码 {idx + 1}/{total}")
                timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                file_path = os.path.join(dir_path, f"qr_{timestamp}_{idx + 1}.png")
                img.save(file_path)
                time.sleep(0.05)  # 避免UI卡顿

            self.log(f"已保存 {len(self.qr_images)} 个二维码到: {dir_path}")
            self.update_progress(100, "保存完成！")
            messagebox.showinfo("成功", f"已保存 {len(self.qr_images)} 个二维码")
        except Exception as e:
            self.log(f"保存失败: {str(e)}")
            self.update_progress(0, f"保存失败: {str(e)}")
            messagebox.showerror("错误", f"保存失败: {str(e)}")
        finally:
            self.update_progress(0, "就绪")

    # 辅助方法
    def decode_qr(self, filepath):
        """解码单个二维码"""
        try:
            img = Image.open(filepath)
            results = pyzbar.decode(img)
            return [r.data.decode('utf-8') for r in results if r.type == 'QRCODE']
        except Exception as e:
            self.log(f"解码失败 {filepath}: {str(e)}")
            return []

    def combine_data(self, chunks):
        """合并分块数据"""
        try:
            # 提取所有分块数据
            chunks_dict = {}
            total_chunks = 0
            current_chunks = 0
            version = 0
            mode = "file"  # 默认文件模式

            # 首先收集所有分块信息
            for chunk in chunks:
                if chunk.startswith("QR:"):
                    # 分块格式: "QR:2/5|v8|mode|base64数据"
                    parts = chunk.split('|', 3)
                    if len(parts) < 4:
                        continue

                    header = parts[0]
                    version_part = parts[1]
                    mode_part = parts[2]
                    data_part = parts[3]

                    # 解析模式
                    mode = mode_part

                    # 解析版本
                    if version_part.startswith("v"):
                        try:
                            version = int(version_part[1:])
                        except:
                            version = 0

                    # 解析分块头: "QR:2/5"
                    chunk_info = header.split(':')[1]
                    chunk_num, total = chunk_info.split('/')

                    chunks_dict[int(chunk_num)] = data_part
                    total_chunks = int(total)
                    current_chunks += 1
                else:
                    # 单个二维码情况
                    return base64.b64decode(chunk)

            # 检查是否收集到所有分块
            if current_chunks != total_chunks:
                missing = [i for i in range(1, total_chunks + 1) if i not in chunks_dict]
                raise ValueError(f"数据不完整: 缺少分块 {missing}")

            # 按顺序组合分块
            combined_b64 = ''.join(chunks_dict[i] for i in sorted(chunks_dict.keys()))
            return base64.b64decode(combined_b64)

        except Exception as e:
            self.log(f"合并数据失败: {str(e)}")
            return None

    def show_qr(self):
        """显示当前二维码"""
        if not self.qr_images:
            # 清除画布
            self.qr_canvas.delete("all")
            return

        # 清除画布
        self.qr_canvas.delete("all")

        # 获取当前二维码图像
        name, img = self.qr_images[self.current_qr_index]

        # 获取画布尺寸
        canvas_width = self.qr_canvas.winfo_width()
        canvas_height = self.qr_canvas.winfo_height()

        if canvas_width <= 1 or canvas_height <= 1:
            return

        # 计算缩放比例
        img_width, img_height = img.size
        width_ratio = canvas_width / img_width
        height_ratio = canvas_height / img_height
        scale_ratio = min(width_ratio, height_ratio, 1) * 0.95  # 留出边距

        # 计算新尺寸
        new_width = int(img_width * scale_ratio)
        new_height = int(img_height * scale_ratio)

        # 调整图像大小
        img_resized = img.resize((new_width, new_height), Image.Resampling.LANCZOS)
        img_tk = ImageTk.PhotoImage(img_resized)

        # 居中显示
        x = (canvas_width - new_width) // 2
        y = (canvas_height - new_height) // 2

        # 保存引用并显示
        self.qr_canvas.image = img_tk
        self.qr_canvas.create_image(x, y, anchor=tk.NW, image=img_tk)

        # 更新位置标签
        self.qr_pos_label.config(text=f"{self.current_qr_index + 1}/{len(self.qr_images)}")
        self.qr_info_label.config(text=f"二维码: {name}")

    def resize_qr(self, event):
        """画布大小变化时重绘二维码"""
        if self.qr_images:
            self.show_qr()

    def prev_qr(self):
        """显示上一个二维码"""
        if self.current_qr_index > 0:
            self.current_qr_index -= 1
            self.show_qr()
            self.update_navigation()

    def next_qr(self):
        """显示下一个二维码"""
        if self.current_qr_index < len(self.qr_images) - 1:
            self.current_qr_index += 1
            self.show_qr()
            self.update_navigation()

    def update_navigation(self):
        """更新导航按钮状态"""
        total = len(self.qr_images)
        self.prev_btn.config(state=tk.NORMAL if self.current_qr_index > 0 else tk.DISABLED)
        self.next_btn.config(state=tk.NORMAL if self.current_qr_index < total - 1 else tk.DISABLED)
        self.qr_pos_label.config(text=f"{self.current_qr_index + 1}/{total}")

    def update_capacity(self, *args):
        """更新容量设置"""
        try:
            self.max_chunk_size = int(self.capacity_var.get())
        except ValueError:
            self.capacity_var.set(str(self.max_chunk_size))

    def browse_file(self):
        """浏览文件"""
        if self.mode == "region":
            filetypes = [("Excel文件", "*.xlsx *.xls")]
        else:
            filetypes = [("所有文件", "*.*")]

        path = filedialog.askopenfilename(
            title="选择文件",
            filetypes=filetypes
        )
        if path:
            self.file_path.set(path)

    def file_changed(self, *args):
        """文件路径变化时加载sheet（仅Excel区域模式）"""
        path = self.file_path.get()
        if path and os.path.exists(path) and self.mode == "region":
            try:
                self.workbook = load_workbook(path)
                self.sheet_names = self.workbook.sheetnames
                self.sheet_combo['values'] = self.sheet_names
                if self.sheet_names:
                    self.sheet_var.set(self.sheet_names[0])
                self.log(f"已加载Excel文件: {os.path.basename(path)}")
            except Exception as e:
                self.log(f"加载Excel失败: {str(e)}")

    def validate_inputs(self):
        """验证输入"""
        if not self.file_path.get():
            messagebox.showwarning("警告", "请选择文件")
            return False

        if not os.path.exists(self.file_path.get()):
            messagebox.showwarning("警告", "文件不存在")
            return False

        if self.mode == "region":
            if not self.region_entry.get():
                messagebox.showwarning("警告", "请输入区域坐标")
                return False

            if not self.sheet_var.get():
                messagebox.showwarning("警告", "请选择Sheet")
                return False

        return True


class QRProcessor:
    def __init__(self, output_dir):
        self.output_dir = output_dir
        os.makedirs(output_dir, exist_ok=True)

    def serialize_excel_region(self, excel_path, region, sheet_name=None, version=8, progress_callback=None):
        """序列化Excel区域"""
        if progress_callback:
            progress_callback(0, "加载Excel文件...")

        wb = load_workbook(excel_path)

        # 选择sheet
        if sheet_name:
            ws = wb[sheet_name]
        else:
            ws = wb.active

        # 解析区域坐标
        min_col, min_row, max_col, max_row = self.parse_region(region)

        # 收集数据
        data = {
            'data': [],
            'styles': [],
            'merged': [m.coord for m in ws.merged_cells.ranges],
            'meta': {
                'source': os.path.basename(excel_path),
                'sheet': sheet_name or ws.title,
                'region': region,
                'version': version,
                'timestamp': datetime.now().isoformat(),
                'mode': 'region'
            }
        }

        total_rows = max_row - min_row + 1
        for row_idx, row in enumerate(ws.iter_rows(min_row=min_row, max_row=max_row,
                                                   min_col=min_col, max_col=max_col)):
            row_data = []
            row_styles = []

            for cell in row:
                # 处理不同类型的数据
                if isinstance(cell.value, str) and len(cell.value) > 1000:
                    row_data.append(cell.value[:1000] + "...[TRUNCATED]")
                else:
                    row_data.append(cell.value)
                row_styles.append(self.get_style(cell))

            data['data'].append(row_data)
            data['styles'].append(row_styles)

            # 更新进度
            if progress_callback and total_rows > 0:
                progress = (row_idx + 1) / total_rows * 50
                progress_callback(progress, f"处理行 {row_idx + 1}/{total_rows}")

        # 序列化并压缩
        if progress_callback:
            progress_callback(60, "序列化数据...")

        serialized = pickle.dumps(data)

        if progress_callback:
            progress_callback(70, "压缩数据...")

        compressed = zlib.compress(serialized)

        # 添加校验和
        if progress_callback:
            progress_callback(80, "添加校验和...")

        checksum = zlib.crc32(compressed)
        compressed_with_checksum = struct.pack(">I", checksum) + compressed

        return compressed_with_checksum

    def serialize_file(self, file_path, version=8, progress_callback=None):
        """序列化任意文件"""
        try:
            if progress_callback:
                progress_callback(0, "读取文件...")

            # 获取文件大小
            file_size = os.path.getsize(file_path)
            if file_size > 10 * 1024 * 1024:  # 10MB
                raise ValueError("文件过大，建议使用区域模式")

            # 读取整个文件
            with open(file_path, 'rb') as f:
                file_data = f.read()

            if progress_callback:
                progress_callback(30, "压缩数据...")

            # 压缩数据
            compressed = zlib.compress(file_data)

            if progress_callback:
                progress_callback(70, "添加校验和...")

            # 添加校验和
            checksum = zlib.crc32(compressed)
            compressed_with_checksum = struct.pack(">I", checksum) + compressed

            # 添加文件模式标记
            file_marker = b"FILE_MODE:"
            final_data = file_marker + compressed_with_checksum

            # 保存副本
            filename = os.path.basename(file_path)
            with open(os.path.join(self.output_dir, f"{filename}.qrdat"), 'wb') as f:
                f.write(final_data)

            return final_data

        except Exception as e:
            raise ValueError(f"文件序列化失败: {str(e)}")

    def restore(self, data, output_path=None):
        """从数据恢复文件"""
        try:
            # 检查是否为文件模式
            if data.startswith(b"FILE_MODE:"):
                return self.restore_file(data[10:], output_path)
            else:
                return self.restore_excel_region(data, output_path)

        except Exception as e:
            # 保存原始数据用于调试
            debug_path = os.path.join(self.output_dir, "restore_debug.dat")
            with open(debug_path, 'wb') as f:
                f.write(data)
            raise ValueError(f"恢复失败: {str(e)}\n原始数据已保存至: {debug_path}")

    def restore_excel_region(self, data, output_path=None):
        """恢复Excel区域数据"""
        # 验证数据完整性
        if len(data) < 4:
            raise ValueError("数据过短，无法恢复")

        # 提取校验和
        stored_checksum = struct.unpack(">I", data[:4])[0]
        actual_data = data[4:]

        # 验证校验和
        actual_checksum = zlib.crc32(actual_data)
        if stored_checksum != actual_checksum:
            raise ValueError(f"数据校验失败: {stored_checksum} != {actual_checksum}")

        # 尝试解压数据
        try:
            decompressed = zlib.decompress(actual_data)
        except zlib.error as e:
            # 尝试不解压直接使用
            try:
                decompressed = actual_data
            except:
                raise ValueError(f"解压失败: {str(e)}")

        # 反序列化
        try:
            restored = pickle.loads(decompressed)
        except pickle.UnpicklingError as e:
            raise ValueError(f"反序列化失败: {str(e)}")

        # 检查模式
        meta = restored.get('meta', {})
        if meta.get('mode') != 'region':
            raise ValueError("数据模式不匹配，请使用文件模式恢复")

        # 检查版本兼容性
        data_version = meta.get('version', 0)
        if data_version < 4:
            raise ValueError(f"不兼容的数据版本: {data_version} (需要4+)")

        # 创建新工作簿
        wb = Workbook()
        ws = wb.active

        # 设置sheet名称
        sheet_name = meta.get('sheet', 'Restored')
        if sheet_name:
            ws.title = sheet_name[:30]  # Excel sheet名称长度限制

        # 恢复数据
        total_rows = len(restored['data'])
        for r, (row_data, row_styles) in enumerate(zip(restored['data'], restored['styles'])):
            for c, (value, style) in enumerate(zip(row_data, row_styles)):
                cell = ws.cell(row=r + 1, column=c + 1, value=value)
                self.apply_style(cell, style)

        # 恢复合并单元格
        for merged in restored.get('merged', []):
            try:
                ws.merge_cells(merged)
            except:
                continue

        # 设置输出路径
        if not output_path:
            source = meta.get('source', 'restored')
            name = os.path.splitext(source)[0]
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            output_path = os.path.join(self.output_dir, f"{name}_{timestamp}_restored.xlsx")

        wb.save(output_path)
        return output_path

    def restore_file(self, data, output_path=None):
        """恢复任意文件"""
        # 验证数据完整性
        if len(data) < 4:
            raise ValueError("数据过短，无法恢复")

        # 提取校验和
        stored_checksum = struct.unpack(">I", data[:4])[0]
        actual_data = data[4:]

        # 验证校验和
        actual_checksum = zlib.crc32(actual_data)
        if stored_checksum != actual_checksum:
            raise ValueError(f"数据校验失败: {stored_checksum} != {actual_checksum}")

        # 解压数据
        try:
            decompressed = zlib.decompress(actual_data)
        except zlib.error as e:
            raise ValueError(f"解压失败: {str(e)}")

        # 设置输出路径
        if not output_path:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            output_path = os.path.join(self.output_dir, f"restored_file_{timestamp}")

        # 保存文件
        with open(output_path, 'wb') as f:
            f.write(decompressed)

        return output_path

    def create_qr_codes(self, data, max_size=1800, version=8, mode="file", progress_callback=None):
        """生成二维码序列"""
        # 计算base64编码后的最大原始数据大小
        max_raw_size = int(max_size * 0.7)  # 考虑base64开销

        # 如果数据很小，直接生成单个二维码
        if len(data) <= max_raw_size:
            # 使用base64编码
            base64_data = base64.b64encode(data).decode('utf-8')
            return [("single", self.create_single_qr(base64_data, f"{mode}"))]

        # 计算需要多少分块
        total_chunks = (len(data) + max_raw_size - 1) // max_raw_size
        chunks = []

        # 大数据分块处理
        for i in range(total_chunks):
            start = i * max_raw_size
            end = min(start + max_raw_size, len(data))
            chunk_data = data[start:end]

            # 添加分块头并使用base64编码
            header = f"QR:{i + 1}/{total_chunks}|v{version}|{mode}|"
            base64_chunk = base64.b64encode(chunk_data).decode('utf-8')

            # 检查总长度
            full_chunk = header + base64_chunk
            if len(full_chunk) > max_size:
                # 如果超出，减小分块大小
                new_max_raw_size = int(max_raw_size * 0.9)
                return self.create_qr_codes(data, new_max_raw_size, version, mode, progress_callback)

            # 创建二维码
            name = f"chunk_{i + 1}_of_{total_chunks}"
            img = self.create_single_qr(full_chunk, f"{i + 1}/{total_chunks}")
            chunks.append((name, img))

            # 更新进度
            if progress_callback:
                progress = (i + 1) / total_chunks * 100
                progress_callback(progress, f"生成二维码 {i + 1}/{total_chunks}")

        return chunks

    def create_single_qr(self, data, counter=None):
        """创建单个二维码"""
        qr = qrcode.QRCode(
            version=None,
            error_correction=qrcode.constants.ERROR_CORRECT_L,
            box_size=10,  # 固定大小
            border=4,
        )
        qr.add_data(data)
        qr.make(fit=True)

        img = qr.make_image(fill_color="black", back_color="white")

        # 添加标记文本
        if counter:
            draw = ImageDraw.Draw(img)
            try:
                font = ImageFont.truetype("arial.ttf", 16)
            except:
                try:
                    font = ImageFont.truetype("Arial.ttf", 16)
                except:
                    font = ImageFont.load_default()

            text = counter
            bbox = draw.textbbox((0, 0), text, font=font)
            text_w = bbox[2] - bbox[0]
            text_h = bbox[3] - bbox[1]

            # 在右下角添加文本
            draw.rectangle(
                [(img.width - text_w - 10, img.height - text_h - 10),
                 (img.width, img.height)],
                fill="white"
            )
            draw.text(
                (img.width - text_w - 5, img.height - text_h - 5),
                text,
                font=font,
                fill="black"
            )

        return img

    # 辅助方法
    def parse_region(self, region):
        """解析区域坐标 - 增强容错性"""
        # 移除空格并转换为大写
        region = region.replace(" ", "").upper()

        if ':' in region:
            start, end = region.split(':', 1)
        else:
            start = end = region

        # 使用正则表达式提取列和行
        pattern = r"([A-Z]+)(\d+)"
        start_match = re.match(pattern, start)
        end_match = re.match(pattern, end)

        if not start_match or not end_match:
            raise ValueError(f"无效的区域格式: {region}")

        start_col = start_match.group(1)
        start_row = int(start_match.group(2))
        end_col = end_match.group(1)
        end_row = int(end_match.group(2))

        # 列字母转数字
        def col_to_num(col):
            num = 0
            for c in col:
                if c.isalpha():
                    num = num * 26 + (ord(c) - ord('A')) + 1
            return num

        return (
            col_to_num(start_col),
            start_row,
            col_to_num(end_col),
            end_row
        )

    def get_style(self, cell):
        """获取单元格样式 - 增强容错性"""
        try:
            return {
                'font': self.copy_font(cell.font) if cell.font else None,
                'fill': self.copy_fill(cell.fill) if cell.fill else None,
                'border': self.copy_border(cell.border) if cell.border else None,
                'alignment': self.copy_alignment(cell.alignment) if cell.alignment else None,
                'format': cell.number_format
            }
        except Exception:
            return {}

    def apply_style(self, cell, style):
        """应用样式到单元格 - 安全处理None值"""
        if style.get('font'):
            try:
                cell.font = style['font']
            except Exception:
                pass
        if style.get('fill'):
            try:
                cell.fill = style['fill']
            except Exception:
                pass
        if style.get('border'):
            try:
                cell.border = style['border']
            except Exception:
                pass
        if style.get('alignment'):
            try:
                cell.alignment = style['alignment']
            except Exception:
                pass
        if style.get('format'):
            try:
                cell.number_format = style['format']
            except Exception:
                pass

    def copy_font(self, font):
        if not font: return None
        return Font(
            name=font.name, size=font.size, bold=font.bold,
            italic=font.italic, strike=font.strike,
            color=font.color
        )

    def copy_fill(self, fill):
        if not fill: return None
        return PatternFill(
            fill_type=fill.fill_type,
            start_color=fill.start_color,
            end_color=fill.end_color
        )

    def copy_border(self, border):
        if not border: return None
        return Border(
            left=border.left, right=border.right,
            top=border.top, bottom=border.bottom
        )

    def copy_alignment(self, alignment):
        if not alignment: return None
        return Alignment(
            horizontal=alignment.horizontal,
            vertical=alignment.vertical,
            wrap_text=alignment.wrap_text,
            shrink_to_fit=alignment.shrink_to_fit,
            indent=alignment.indent
        )


if __name__ == "__main__":
    root = tk.Tk()
    app = FileQRApp(root)
    root.mainloop()